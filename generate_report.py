#!/usr/bin/env python3
"""
KOCSKIN Weekly Report Generator (v3)
=====================================
Pipeline
--------
1. Read 2 source files from Google Drive:
     - 商店銷售統計表*.xlsx  (91APP 每日訂單)
     - 商品報表*.csv          (91APP 商品明細)
2. If either is NOT newer than last run → skip (exit 3).
3. Use the date range from the sales xlsx as [since, until].
4. Pull FB Marketing API for that date range → ad insights (was CSV upload).
5. Pull GA4 Data API for that date range → traffic analytics (new tab).
6. Parse + strip sensitive (成本 / 毛利 / 利潤) columns.
7. Merge into accumulated history.json.
8. Render dashboard HTML → encrypt via staticrypt → publish to /dist.

Designed to run inside GitHub Actions on a daily cron schedule.
"""

from __future__ import annotations

import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

# Our API clients
from fb_api_client import (
    check_token_expiry,
    warn_if_expiring,
    fetch_ad_insights,
    FbApiError,
)
from ga_api_client import fetch_ga_report

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).parent.resolve()
TEMPLATE_DIR = ROOT / "templates"
STATE_DIR = ROOT / "state"
LOG_DIR = ROOT / "logs"
DIST_DIR = ROOT / "dist"
DOWNLOAD_DIR = ROOT / "downloads"

HISTORY_FILE = STATE_DIR / "history.json"
LAST_RUN_FILE = STATE_DIR / "last_run.json"

# Only 2 file patterns now — FB/GA come from APIs
PATTERNS = {
    "sales_xlsx": re.compile(r"^商店銷售統計表.*\.xlsx$", re.IGNORECASE),
    "product_csv": re.compile(r"^商品報表.*\.csv$", re.IGNORECASE),
}

# Columns we MUST strip from sales xlsx (never shown to employees)
SENSITIVE_SALES_COLS = {"成本", "毛利率", "毛利", "毛利金額"}

# Columns we MUST strip from product csv (if present)
SENSITIVE_PRODUCT_COLS = {"成本", "毛利", "毛利率", "毛利金額", "利潤"}

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

LOG_DIR.mkdir(parents=True, exist_ok=True)
log_file = LOG_DIR / f"{datetime.now().strftime('%Y%m')}.log"

logger = logging.getLogger("kocskin_report")
logger.setLevel(logging.INFO)

_fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

_fh = logging.FileHandler(log_file, encoding="utf-8")
_fh.setFormatter(_fmt)
logger.addHandler(_fh)

_sh = logging.StreamHandler(sys.stdout)
_sh.setFormatter(_fmt)
logger.addHandler(_sh)

# Propagate into child loggers (fb_api_client / ga_api_client)
for child in ("kocskin_report.fb", "kocskin_report.ga"):
    logging.getLogger(child).setLevel(logging.INFO)


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        logger.warning("Failed to read %s: %s — using default", path, exc)
        return default


def _save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    tmp.replace(path)


def _nowiso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# Google Drive
# ---------------------------------------------------------------------------

def build_drive_client(sa_info: Dict[str, Any]):
    creds = service_account.Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def find_drive_folder_id(drive, folder_name: str) -> str:
    q = (
        f"name = '{folder_name}' and "
        f"mimeType = 'application/vnd.google-apps.folder' and "
        f"trashed = false"
    )
    res = drive.files().list(q=q, fields="files(id,name)", pageSize=10).execute()
    files = res.get("files", [])
    if not files:
        raise RuntimeError(
            f"Google Drive folder '{folder_name}' not found. "
            f"Did you share it with the service account email?"
        )
    if len(files) > 1:
        logger.warning("Multiple folders named '%s' found — using first", folder_name)
    return files[0]["id"]


def list_drive_files(drive, folder_id: str) -> List[Dict]:
    files: List[Dict] = []
    page_token: Optional[str] = None
    while True:
        res = drive.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, mimeType, modifiedTime, size)",
            pageSize=100,
            pageToken=page_token,
        ).execute()
        files.extend(res.get("files", []))
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return files


def download_drive_file(drive, file_id: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    request = drive.files().get_media(fileId=file_id)
    with dest.open("wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _status, done = downloader.next_chunk()


def match_patterns(files: List[Dict]) -> Dict[str, Dict]:
    matched: Dict[str, Dict] = {}
    for key, pattern in PATTERNS.items():
        candidates = [f for f in files if pattern.match(f["name"])]
        if not candidates:
            continue
        candidates.sort(key=lambda x: x["modifiedTime"], reverse=True)
        matched[key] = candidates[0]
    return matched


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_sales_xlsx(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    keep_idx = [i for i, h in enumerate(header) if h and h not in SENSITIVE_SALES_COLS]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(v is None for v in r):
            continue
        rec: Dict[str, Any] = {}
        for i in keep_idx:
            key = header[i]
            val = r[i] if i < len(r) else None
            if key == "日期" and val is not None:
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val)[:10]
            rec[key] = val
        date_val = rec.get("日期")
        if date_val and re.match(r"^\d{4}-\d{2}-\d{2}$", str(date_val)):
            out.append(rec)
    logger.info("Parsed sales xlsx: %d daily rows", len(out))
    return out


def parse_product_csv(path: Path) -> List[Dict[str, Any]]:
    df = pd.read_csv(path, encoding="utf-8-sig")
    drop_cols = [c for c in df.columns if c in SENSITIVE_PRODUCT_COLS]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    for c in df.columns:
        if c in ("已售商品數量", "訂單數(TS)", "訂單金額(TS)", "售價金額(TS)"):
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    records = df.to_dict(orient="records")
    logger.info("Parsed product csv: %d rows", len(records))
    return records


def scrub_ads_sensitive(ads: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Drop any key containing 成本/毛利/利潤 (defensive, since API fields are fixed)."""
    bad_keys = {"成本", "毛利", "利潤"}
    cleaned: List[Dict[str, Any]] = []
    for a in ads:
        cleaned.append({k: v for k, v in a.items()
                        if not any(w in str(k) for w in bad_keys)})
    return cleaned


# ---------------------------------------------------------------------------
# History merge
# ---------------------------------------------------------------------------

def merge_history(
    history: Dict[str, Any],
    sales: List[Dict],
    products: List[Dict],
    ads: List[Dict],
    ga: Optional[Dict[str, Any]],
    run_date: str,
) -> Dict[str, Any]:
    # daily_sales: merge by 日期
    existing = {row["日期"]: row for row in history.get("daily_sales", []) if row.get("日期")}
    for row in sales:
        existing[row["日期"]] = row
    history["daily_sales"] = sorted(existing.values(), key=lambda r: r["日期"])

    # product_snapshots
    prod_snaps = history.get("product_snapshots", [])
    prod_snaps = [s for s in prod_snaps if s.get("run_date") != run_date]
    prod_snaps.append({"run_date": run_date, "items": products})
    prod_snaps.sort(key=lambda s: s["run_date"])
    history["product_snapshots"] = prod_snaps[-52:]

    # ad_snapshots
    ad_snaps = history.get("ad_snapshots", [])
    ad_snaps = [s for s in ad_snaps if s.get("run_date") != run_date]
    ad_snaps.append({"run_date": run_date, "items": ads})
    ad_snaps.sort(key=lambda s: s["run_date"])
    history["ad_snapshots"] = ad_snaps[-52:]

    # ga_snapshots (only store latest — GA data is derived, not cumulative from run-to-run)
    if ga is not None:
        ga_snaps = history.get("ga_snapshots", [])
        ga_snaps = [s for s in ga_snaps if s.get("run_date") != run_date]
        ga_snaps.append({"run_date": run_date, "data": ga})
        ga_snaps.sort(key=lambda s: s["run_date"])
        history["ga_snapshots"] = ga_snaps[-52:]

    runs = history.get("runs", [])
    runs.append({"run_date": run_date, "at": _nowiso()})
    history["runs"] = runs[-200:]
    return history


# ---------------------------------------------------------------------------
# Aggregations for HTML
# ---------------------------------------------------------------------------

def _num(v, default=0):
    try:
        if v is None or v == "":
            return default
        f = float(v)
        if f != f:  # NaN
            return default
        return f
    except (TypeError, ValueError):
        return default


def build_dashboard_payload(history: Dict[str, Any]) -> Dict[str, Any]:
    sales = history.get("daily_sales", [])
    products_latest = (history.get("product_snapshots") or [{}])[-1].get("items", [])
    ads_latest = (history.get("ad_snapshots") or [{}])[-1].get("items", [])
    ga_latest = None
    ga_snaps = history.get("ga_snapshots") or []
    if ga_snaps:
        ga_latest = ga_snaps[-1].get("data")

    # ----- Daily series -----
    daily_series = []
    for r in sales:
        daily_series.append({
            "date": r.get("日期"),
            "orders": _num(r.get("訂單筆數(TS)")),
            "items": _num(r.get("商品數")),
            "gross_amount": _num(r.get("訂單金額(折扣後)")),
            "discount": _num(r.get("折扣金額")),
            "cancel_orders": _num(r.get("取消單筆數")),
            "cancel_amount": _num(r.get("取消單金額")),
            "return_orders": _num(r.get("退貨結案 退貨單筆數")),
            "return_amount": _num(r.get("退貨結案 退貨金額")),
            "net_sales": _num(r.get("淨銷售額")),
        })

    # ----- Sales KPIs -----
    total_orders = sum(d["orders"] for d in daily_series)
    total_gross = sum(d["gross_amount"] for d in daily_series)
    total_discount = sum(d["discount"] for d in daily_series)
    total_net = sum(d["net_sales"] for d in daily_series)
    total_items = sum(d["items"] for d in daily_series)
    total_cancel_orders = sum(d["cancel_orders"] for d in daily_series)
    total_cancel_amount = sum(d["cancel_amount"] for d in daily_series)
    total_return = sum(d["return_orders"] for d in daily_series)
    cancel_rate = (total_cancel_orders / total_orders * 100) if total_orders else 0
    return_rate = (total_return / total_orders * 100) if total_orders else 0
    avg_order_value = (total_gross / total_orders) if total_orders else 0

    sales_dates = sorted([d["date"] for d in daily_series if d["date"]])
    period_start = sales_dates[0] if sales_dates else "-"
    period_end = sales_dates[-1] if sales_dates else "-"
    period_days = len(sales_dates) if sales_dates else 1

    # ----- Top products -----
    top_products = []
    for p in products_latest:
        name_raw = p.get("商品頁名稱")
        sid = p.get("商品頁序號")
        if name_raw is None or str(name_raw).strip().lower() in ("", "nan", "none", "null"):
            continue
        if sid is None or str(sid).strip().lower() in ("", "nan", "none", "null"):
            continue
        qty = _num(p.get("已售商品數量"))
        if qty <= 0:
            continue
        top_products.append({
            "name": str(name_raw)[:60],
            "qty": qty,
            "orders": _num(p.get("訂單數(TS)")),
            "amount": _num(p.get("訂單金額(TS)")),
        })
    top_products.sort(key=lambda x: x["qty"], reverse=True)
    top_products = top_products[:15]

    # ----- Ads aggregate by 廣告組合名稱 -----
    ad_groups: Dict[str, Dict[str, float]] = {}
    for a in ads_latest:
        g = str(a.get("廣告組合名稱") or "未分組")
        d = ad_groups.setdefault(g, {
            "spend": 0.0, "impressions": 0.0, "reach": 0.0,
            "purchases": 0.0, "roas_weighted": 0.0, "ad_count": 0,
            "has_active": False,
        })
        spend = _num(a.get("花費金額 (TWD)"))
        d["spend"] += spend
        d["impressions"] += _num(a.get("曝光次數"))
        d["reach"] += _num(a.get("觸及人數"))
        d["purchases"] += _num(a.get("購買次數"))
        roas = _num(a.get("購買 ROAS（廣告投資報酬率）"))
        d["roas_weighted"] += roas * spend
        d["ad_count"] += 1
        if a.get("廣告投遞") == "active":
            d["has_active"] = True

    ad_group_rows = []
    for g, d in ad_groups.items():
        avg_roas = (d["roas_weighted"] / d["spend"]) if d["spend"] else 0
        ad_group_rows.append({
            "group": g,
            "status": "active" if d["has_active"] else "inactive",
            "spend": round(d["spend"]),
            "impressions": int(d["impressions"]),
            "reach": int(d["reach"]),
            "purchases": int(d["purchases"]),
            "roas": round(avg_roas, 2),
            "cpm": round(d["spend"] / d["impressions"] * 1000, 2) if d["impressions"] else 0,
            "ad_count": d["ad_count"],
        })
    ad_group_rows.sort(key=lambda x: x["spend"], reverse=True)

    # ----- Ad totals -----
    total_spend = sum(_num(a.get("花費金額 (TWD)")) for a in ads_latest)
    total_purchases = sum(_num(a.get("購買次數")) for a in ads_latest)
    total_impressions = sum(_num(a.get("曝光次數")) for a in ads_latest)
    total_reach = sum(_num(a.get("觸及人數")) for a in ads_latest)

    roas_weighted_sum = 0.0
    for a in ads_latest:
        spend = _num(a.get("花費金額 (TWD)"))
        roas = _num(a.get("購買 ROAS（廣告投資報酬率）"))
        roas_weighted_sum += spend * roas
    meta_roas = (roas_weighted_sum / total_spend) if total_spend else 0
    meta_revenue = meta_roas * total_spend
    blended_roas = (total_gross / total_spend) if total_spend else 0
    cpa = (total_spend / total_purchases) if total_purchases else 0
    meta_share_of_orders = (total_purchases / total_orders * 100) if total_orders else 0
    meta_attribution_rate = (meta_revenue / total_gross * 100) if total_gross else 0
    avg_daily_spend = (total_spend / period_days) if period_days else 0
    overall_cpm = (total_spend / total_impressions * 1000) if total_impressions else 0

    def _bucket(status_key: str):
        items = [a for a in ads_latest
                 if str(a.get("廣告投遞") or "").strip().lower() == status_key]
        count = len(items)
        spend = sum(_num(a.get("花費金額 (TWD)")) for a in items)
        purchases = sum(_num(a.get("購買次數")) for a in items)
        rw = sum(_num(a.get("花費金額 (TWD)")) * _num(a.get("購買 ROAS（廣告投資報酬率）"))
                 for a in items)
        roas = (rw / spend) if spend else 0
        return {
            "count": count,
            "spend": int(round(spend)),
            "purchases": int(purchases),
            "roas": round(roas, 2),
            "spend_share": round(spend / total_spend * 100, 1) if total_spend else 0,
        }

    active_bucket = _bucket("active")
    inactive_bucket = _bucket("inactive")

    # ----- GA block -----
    ga_block = None
    if ga_latest:
        ga_block = {
            "daily": ga_latest.get("daily", []),
            "traffic_sources": ga_latest.get("traffic_sources", []),
            "devices": ga_latest.get("devices", []),
            "landing_pages": ga_latest.get("landing_pages", []),
            "totals": ga_latest.get("totals", {}),
        }

    return {
        "generated_at": _nowiso(),
        "period_start": period_start,
        "period_end": period_end,
        "period_days": period_days,
        "kpis_sales": {
            "total_orders": int(total_orders),
            "total_gross": int(total_gross),
            "total_net_sales": int(total_net),
            "total_items": int(total_items),
            "avg_order_value": int(avg_order_value),
            "total_discount": int(total_discount),
            "total_cancel_orders": int(total_cancel_orders),
            "total_cancel_amount": int(total_cancel_amount),
            "cancel_rate": round(cancel_rate, 2),
            "return_rate": round(return_rate, 2),
        },
        "kpis_ad": {
            "total_spend": int(round(total_spend)),
            "avg_daily_spend": int(round(avg_daily_spend)),
            "blended_roas": round(blended_roas, 2),
            "meta_roas": round(meta_roas, 2),
            "meta_revenue": int(round(meta_revenue)),
            "meta_purchases": int(total_purchases),
            "cpa": int(round(cpa)),
            "meta_share_of_orders": round(meta_share_of_orders, 1),
            "total_impressions": int(total_impressions),
            "total_reach": int(total_reach),
            "overall_cpm": int(round(overall_cpm)),
            "meta_attribution_rate": round(meta_attribution_rate, 1),
            "active": active_bucket,
            "inactive": inactive_bucket,
        },
        "daily_series": daily_series,
        "top_products": top_products,
        "ad_groups": ad_group_rows,
        "ga": ga_block,
        "runs_count": len(history.get("runs", [])),
    }


# ---------------------------------------------------------------------------
# Render + encrypt
# ---------------------------------------------------------------------------

def render_html(payload: Dict[str, Any]) -> str:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATE_DIR)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    tpl = env.get_template("report_template.html")
    return tpl.render(
        payload=payload,
        payload_json=json.dumps(payload, ensure_ascii=False),
    )


def encrypt_with_staticrypt(html: str, password: str, out_path: Path) -> None:
    with tempfile.TemporaryDirectory() as td:
        src = Path(td) / "index.html"
        src.write_text(html, encoding="utf-8")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        template_path = Path(__file__).parent / "templates" / "password_template.html"
        cmd = [
            "npx", "--yes", "staticrypt",
            str(src),
            "-p", password,
            "--short",
            "-d", str(out_path.parent),
            "-t", str(template_path),
            "--template-title", "KOCSKIN Reports",
            "--template-instructions", "請輸入內部密碼查看報表",
            "--template-placeholder", "密碼",
            "--template-button", "查看報表",
        ]
        logger.info("Running staticrypt encryption…")
        res = subprocess.run(cmd, capture_output=True, text=True)
        if res.returncode != 0:
            logger.error("staticrypt stdout: %s", res.stdout)
            logger.error("staticrypt stderr: %s", res.stderr)
            raise RuntimeError("staticrypt failed")
        produced = out_path.parent / "index.html"
        if produced.exists() and produced.resolve() != out_path.resolve():
            shutil.move(str(produced), str(out_path))
        logger.info("Encrypted HTML written to %s", out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(
    drive_folder_name: str,
    password: str,
    skip_encrypt: bool = False,
    local_only: bool = False,
) -> int:
    logger.info("=" * 60)
    logger.info("Run started at %s", _nowiso())

    last_run = _load_json(LAST_RUN_FILE, {
        "last_mtimes": {"sales_xlsx": None, "product_csv": None},
        "last_success": None,
    })
    history = _load_json(HISTORY_FILE, {
        "daily_sales": [], "product_snapshots": [],
        "ad_snapshots": [], "ga_snapshots": [], "runs": [],
    })

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    # ---- 1. Load service account JSON (Drive + GA) ----
    sa_raw = os.environ.get("SERVICE_ACCOUNT_JSON")
    if not sa_raw:
        raise RuntimeError("SERVICE_ACCOUNT_JSON env var is empty")
    sa_info = json.loads(sa_raw)

    # ---- 2. File detection ----
    if local_only:
        logger.info("LOCAL mode — reading from %s", DOWNLOAD_DIR)
        local_files = list(DOWNLOAD_DIR.iterdir())
        matched: Dict[str, Dict] = {}
        for key, pattern in PATTERNS.items():
            cand = [p for p in local_files if pattern.match(p.name)]
            if cand:
                cand.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                p = cand[0]
                matched[key] = {
                    "id": str(p),
                    "name": p.name,
                    "modifiedTime": datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).isoformat(),
                    "_local_path": p,
                }
        drive = None
    else:
        drive = build_drive_client(sa_info)
        folder_id = find_drive_folder_id(drive, drive_folder_name)
        logger.info("Drive folder id: %s", folder_id)
        files = list_drive_files(drive, folder_id)
        logger.info("Folder has %d files", len(files))
        matched = match_patterns(files)

    missing = [k for k in PATTERNS if k not in matched]
    if missing:
        logger.warning("Missing file patterns: %s — aborting", missing)
        logger.info("Matched: %s", {k: v["name"] for k, v in matched.items()})
        return 2

    logger.info("Matched files:")
    for k, v in matched.items():
        logger.info("  %s: %s (mtime=%s)", k, v["name"], v["modifiedTime"])

    prev = last_run.get("last_mtimes", {})
    all_newer = True
    for key, meta in matched.items():
        prev_m = prev.get(key)
        if prev_m and meta["modifiedTime"] <= prev_m:
            logger.info("  %s not newer (prev=%s, now=%s)", key, prev_m, meta["modifiedTime"])
            all_newer = False
    if not all_newer:
        logger.info("Sales xlsx / product csv not both newer than last run — skipping")
        return 3

    # ---- 3. Download sales + product ----
    local_paths: Dict[str, Path] = {}
    for key, meta in matched.items():
        if local_only:
            local_paths[key] = meta["_local_path"]
        else:
            dest = DOWNLOAD_DIR / meta["name"]
            logger.info("Downloading %s → %s", meta["name"], dest)
            download_drive_file(drive, meta["id"], dest)
            local_paths[key] = dest

    # ---- 4. Parse sales xlsx → get date range ----
    sales = parse_sales_xlsx(local_paths["sales_xlsx"])
    sales_dates = sorted([r.get("日期") for r in sales if r.get("日期")])
    if not sales_dates:
        logger.error("Sales xlsx has no valid dates — cannot query APIs")
        return 1
    since, until = sales_dates[0], sales_dates[-1]
    logger.info("Sales period: %s ~ %s (%d days)", since, until, len(sales_dates))

    products = parse_product_csv(local_paths["product_csv"])

    # ---- 5. FB API ----
    fb_token = os.environ.get("FB_ACCESS_TOKEN", "").strip()
    fb_account = os.environ.get("FB_AD_ACCOUNT_ID", "").strip()
    ads: List[Dict[str, Any]] = []
    if not fb_token or not fb_account:
        logger.error("FB_ACCESS_TOKEN / FB_AD_ACCOUNT_ID not set — skipping FB fetch")
    else:
        try:
            token_info = check_token_expiry(fb_token)
            warn_if_expiring(token_info, threshold_days=10)
            if not token_info.get("is_valid"):
                logger.error("FB token invalid — aborting FB fetch")
            else:
                ads = fetch_ad_insights(fb_token, fb_account, since, until)
                ads = scrub_ads_sensitive(ads)
                logger.info("FB: %d ad records fetched", len(ads))
        except FbApiError as e:
            logger.error("FB API error: %s", e)
        except Exception as e:
            logger.exception("FB unexpected error: %s", e)

    # ---- 6. GA API (optional) ----
    ga_property = os.environ.get("GA4_PROPERTY_ID", "").strip()
    ga: Optional[Dict[str, Any]] = None
    if not ga_property:
        logger.info("GA4_PROPERTY_ID not set — skipping GA fetch")
    else:
        try:
            ga = fetch_ga_report(sa_info, ga_property, since, until)
        except Exception as e:
            logger.exception("GA unexpected error: %s", e)

    # ---- 7. History merge ----
    run_date = datetime.now().strftime("%Y-%m-%d")
    history = merge_history(history, sales, products, ads, ga, run_date)

    # ---- 8. Render + encrypt ----
    payload = build_dashboard_payload(history)
    html = render_html(payload)

    DIST_DIR.mkdir(parents=True, exist_ok=True)
    out_html = DIST_DIR / "index.html"

    if skip_encrypt:
        out_html.write_text(html, encoding="utf-8")
        logger.info("Plain HTML written to %s (encryption skipped)", out_html)
    else:
        encrypt_with_staticrypt(html, password, out_html)

    # ---- 9. Persist state ----
    _save_json(HISTORY_FILE, history)
    last_run["last_mtimes"] = {k: v["modifiedTime"] for k, v in matched.items()}
    last_run["last_success"] = _nowiso()
    _save_json(LAST_RUN_FILE, last_run)

    logger.info(
        "Run complete. orders=%s ad_spend=%s blended_roas=%s ga_sessions=%s",
        payload["kpis_sales"]["total_orders"],
        payload["kpis_ad"]["total_spend"],
        payload["kpis_ad"]["blended_roas"],
        (payload.get("ga") or {}).get("totals", {}).get("sessions", "-"),
    )
    return 0


def main() -> int:
    folder = os.environ.get("DRIVE_FOLDER_NAME", "KOCSKIN_週報")
    password = os.environ.get("REPORT_PASSWORD", "")
    skip_encrypt = os.environ.get("SKIP_ENCRYPT", "").lower() in ("1", "true", "yes")
    local_only = os.environ.get("LOCAL_ONLY", "").lower() in ("1", "true", "yes")

    if not skip_encrypt and not password:
        logger.error("REPORT_PASSWORD is required unless SKIP_ENCRYPT=1")
        return 1

    try:
        code = run(folder, password, skip_encrypt=skip_encrypt, local_only=local_only)
        logger.info("Exit code: %d", code)
        return code
    except Exception as exc:
        logger.exception("Run failed: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
